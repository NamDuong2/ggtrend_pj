#!/usr/bin/env python
# coding: utf-8

# In[41]:


while True:
    print('Nhập chức năng bạn muốn thực hiện:  ')
    print(' ')
    print('1. Lấy dữ liệu trending từ file.') 
    print('2. Xuất báo cáo top 10 trending')
    print('3. Xuất báo cáo search keyword in 2020')
    print('4. Vẽ biểu đồ line chart top 5 trending các từ khóa tìm kiếm nhiều nhất 2020')
    print('5. Vẽ biểu đồ bar chart top 5 trending các từ khóa tìm kiếm nhiều nhất 2019')
    print('6. Thống kê tìm kiếm top trending 5 từ khóa trong 2 năm 2020, 2019')
    print('...')
    print('99. Thoát')
    user = input('Nhập chức năng bạn muốn thực hiện:  ')
    if user == '1':
        import sys
        import time
        import pandas as pd 
        from pytrends.request import TrendReq
        pytrend = TrendReq(hl='VN', tz=360)
        
        #Nhập từ người dùng

        excel_file = input('Nhập tên file excel key trending: ')
        time_range = input('Nhập khoảng thời gian muốn lấy dữ liệu: ')
        import os.path
        if os.path.isfile(excel_file+'.xlsx') == True: # kiem tra file co ton tai khong
            # đọc tự động file excel chứa keyword yêu cầu
            df = pd.read_excel(excel_file+'.xlsx').dropna()

            # tạo list các list chứa keyword của mỗi mục

            all_kw = []
            kw_list = []

            for j in range(0,len(df.columns)):
                for i in range(0,len(df)):        
                    kw_list.append(df.iloc[i,j])        
                    kw_list = [x for x in kw_list if str(x) != 'nan']
                all_kw.append(kw_list)
                kw_list = []

            print(all_kw)
            print('Waiting...')


            import sqlalchemy
            engine = sqlalchemy.create_engine('postgresql://postgres:1234@localhost:5432/postgres')
            con = engine.connect()


            try:
                keywords = []       #list tạm thời
                d = {}
                def check_trends():
                    pytrend.build_payload(keywords, cat=0, timeframe=time_range, geo='VN',gprop='')
                    d[i] = pytrend.interest_over_time()  
                    time.sleep(12)
                    if not d[i].empty:
                        d0 = d[i].drop('isPartial', axis=1).reset_index()
                        d0.rename(columns={str(d0.columns[1]):'value'}, inplace = True)  #đổi tên cột thành tên value
                        d0.insert(0, "keyword", keywords[0], True)    #thêm cột keyword trước cột date
                        if k == 0:
                            d0['trend_type'] = 'news'   #thêm cột trend_type với giá trị news
                        elif k == 1:
                            d0['trend_type'] = 'person'
                        elif k == 2:
                            d0['trend_type'] = 'film'
                        elif k == 3:
                            d0['trend_type'] = 'elearning online'
                        elif k == 4:
                            d0['trend_type'] = 'diseases'
                        elif k == 5:
                            d0['trend_type'] = 'songs'
                        elif k == 6:
                            d0['trend_type'] = 'travel'
                        d0.to_sql('vn_trending', con, if_exists = 'append', chunksize = 100000, index = False)

                i = 0
                for k in range(7):
                    for kw in all_kw[k]:
                        keywords.append(kw)
                        check_trends()
                        keywords.pop()           # mỗi vòng lặp xóa kí tự cuối để có thể chạy liên tục mà không bị giới hạn 5 tìm kiếm của ggtrend
                        i += 1
            except:
                print("Có ngoại lệ ",sys.exc_info()[0]," xảy ra.")
            else:
                print('Đã insert dữ liệu lấy được vào cơ sở dữ liệu thành công!!!')
        
        elif os.path.isfile(excel_file+'.xlsx') == False:
            print('File không tồn tại!!!')
       
    elif user == '2':
        import psycopg2
        import pandas as pd
        import numpy as np
        conn = psycopg2.connect("host=localhost dbname=postgres user=postgres password=1234")

        path = r'C:\Users\ADMIN\result_user_2.xlsx'
        writer = pd.ExcelWriter(path, engine = 'xlsxwriter')
        
        import xlsxwriter
        # Create an new Excel file and add a worksheet.
        workbook = xlsxwriter.Workbook('result_user_2.xlsx')
        worksheet = workbook.add_worksheet()

        # Increase the cell size of the merged cells to highlight the formatting.
        worksheet.set_column('A:D', 12)
        worksheet.set_row(3)



        # Create a format to use in the merged range.
        merge_format = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'})


        # Merge cells.
        worksheet.merge_range('A1:D1', 'DANH SÁCH TỪ KHÓA TÌM KIẾM NHIỀU NHẤT VIỆT NAM', merge_format)
        workbook.close()
        #tạo dataframe 
        df10 = pd.read_sql_query('''
                    SELECT
                        keyword,
                        (SUM(value)) as sum   
                    FROM vn_trending
                    GROUP BY
                        keyword
                    ORDER BY sum DESC
                    FETCH FIRST 10 ROWS ONLY;''', conn)
        df10 = pd.DataFrame(df10)
        
        sql_query2 = {}
        for i in range(0,10):
            sql_query2[i] = pd.read_sql_query(''' 
                    SELECT month_year 
                    FROM
                    (SELECT
                        to_char(date,'MM-YYYY') as month_year,
                        (SUM(value)) as sum,
                        keyword
                    FROM vn_trending
                    WHERE keyword in ('{}')
                    GROUP BY
                        keyword,
                        month_year
                    ORDER BY sum DESC) as foo
                    FETCH first 1 rows only;'''.format(df10.iloc[i,0]),conn)
            top_month = pd.concat(sql_query2,ignore_index=True)      
            top_keytrend = pd.concat([df10,top_month],axis=1)
            top_keytrend.insert(0,'STT',[i for i in range(1,len(df10)+1)],True)        
            top_keytrend.columns = ['STT','keyword','Số lần tìm kiếm','Tháng tìm kiếm nhiều nhất']  
            #lưu  top key trend vào file có sẵn
            import pandas as pd
            import openpyxl
            from openpyxl import load_workbook
            writer = pd.ExcelWriter('result_user_2.xlsx', engine='openpyxl')
            # try to open an existing workbook
            writer.book = load_workbook('result_user_2.xlsx')
            # copy existing sheets
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            # read existing file
            reader = pd.read_excel(r'result_user_2.xlsx')
            # write out the new sheet
            top_keytrend.to_excel(writer,index=False,header=True,startrow=len(reader)+1)
        writer.close()
        print('Đã xuất file result_user_2.xlsx thành công!!!')
        continue
    elif user == '3':
        import psycopg2
        import pandas as pd
        import numpy as np
        conn = psycopg2.connect("host=localhost dbname=postgres user=postgres password=1234")

        path = r'C:\Users\ADMIN\result_user_3.xlsx'
        writer = pd.ExcelWriter(path, engine = 'xlsxwriter')
        
        
        # tạo bảng truy vấn dữ liệu năm 2020
        l=[0,1,5,6,7,8,9,10,11,12,2,3,4] 
        cat = ['news','person','film','elearning online','diseases','songs','travel']
        for name in cat:
            query =  pd.read_sql_query(''' 
                                select *
                                from
                                (SELECT
                                    to_char(date,'MM-YYYY') as myear,
                                    SUM(value) as sum,
                                    keyword
                                FROM vn_trending
                                WHERE trend_type = '{}'
                                GROUP BY
                                    keyword,
                                    to_char(date,'MM-YYYY')
                                ORDER BY
                                    myear ASC) as foo
                                WHERE myear in ('01-2020','02-2020','03-2020','04-2020','05-2020','06-2020','07-2020',
                                                '08-2020','09-2020','10-2020','11-2020','12-2020');'''.format(name), conn)
            
            df = pd.DataFrame(query)
            h = ['0{}-2020'.format(i) for i in range(1,10)]
            h.extend(['{}-2020'.format(i) for i in range(10,13)])
            df.myear.replace(h,['Tháng '+ str(i) for i in range(1,13)],inplace=True)
            df = df.pivot(index='keyword',columns='myear', values='sum')
            df = df.reset_index()   
            df = df[[df.columns[i] for i in l]]
            df.insert(0,'STT',[i for i in range(1,len(df)+1)],True)
            df.to_excel(writer, index = False, sheet_name = '{}'.format(name))
        writer.save()
            
            
        print('Đã xuất file result_user_3.xlsx thành công!!!')
        continue
    elif user == '4':
        import psycopg2
        import pandas as pd
        import numpy as np
        conn = psycopg2.connect("host=localhost dbname=postgres user=postgres password=1234")

        sql4 =  pd.read_sql_query('''
                    SELECT
                        keyword,
                        SUM(value) as sum
                    FROM vn_trending
                    WHERE to_char(date,'MM-YYYY') in ('01-2020','02-2020','03-2020','04-2020','05-2020','06-2020','07-2020',
                                                '08-2020','09-2020','10-2020','11-2020','12-2020')
                    GROUP BY
                        keyword
                    ORDER BY
                        sum DESC
                    FETCH FIRST 5 ROWS ONLY;''',conn)
        top20 = pd.DataFrame(sql4)
        
        # vẽ biểu đồ đường 2020
        import matplotlib.pyplot as plt
        fig = plt.figure(figsize=(10,6))
        plt.plot(top20['keyword'], top20['sum'], color='blue')
        plt.title('TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM 2020', fontsize=14)
        plt.grid(True, axis = 'y')
        plt.show()
        continue
    elif user == '5':
        import psycopg2
        import pandas as pd
        import numpy as np
        conn = psycopg2.connect("host=localhost dbname=postgres user=postgres password=1234")

        sql5 =  pd.read_sql_query('''
            SELECT
                keyword,
                SUM(value) as sum
            FROM vn_trending
            WHERE to_char(date,'MM-YYYY') in ('01-2019','02-2019','03-2019','04-2019','05-2019','06-2019','07-2019',
                                                '08-2019','09-2019','10-2019','11-2019','12-2019')
            GROUP BY
                keyword
            ORDER BY
                sum DESC
            FETCH FIRST 5 ROWS ONLY;''',conn)
        top19 = pd.DataFrame(sql5)


        import matplotlib.pyplot as plt
        fig = plt.figure(figsize=(10,6))
        plt.bar(top19['keyword'], top19['sum'], width = 0.4)
        plt.title('TỪ KHÓA TÌM KIẾM NHIỀU NHẤT TẠI VIỆT NAM 2019', fontsize=14)
        plt.grid(True, axis = 'y')
        plt.show()
        continue
    elif user == '6':
        # merge cell
        import xlsxwriter


        # Create an new Excel file and add a worksheet.
        workbook = xlsxwriter.Workbook('result_user_6.xlsx')
        worksheet = workbook.add_worksheet()

        # Increase the cell size of the merged cells to highlight the formatting.
        worksheet.set_column('A:G', 12)
        worksheet.set_row(3)



        # Create a format to use in the merged range.
        merge_format = workbook.add_format({
            'bold': 1,
            'border': 1,
            'align': 'center',
            'valign': 'vcenter'})


        # Merge 3 cells.
        worksheet.merge_range('A1:G1', 'THỐNG KÊ TÌM KIẾM NHIỀU NHẤT TRONG 2 NĂM', merge_format)
        worksheet.merge_range('A2:D2', 'Năm 2020', merge_format)
        worksheet.merge_range('E2:G2', 'Năm 2019', merge_format)
        workbook.close()
        #top5 tu khoa nhieu nhat 2020
        import psycopg2
        import pandas as pd
        import numpy as np
        conn = psycopg2.connect("host=localhost dbname=postgres user=postgres password=1234")

        path = r'C:\Users\ADMIN\result_user_6.xlsx'
        writer = pd.ExcelWriter(path, engine = 'xlsxwriter')
        
        df5_20 = pd.read_sql_query('''
                    SELECT
                        keyword,
                        (SUM(value)) as sum
                        
                    FROM vn_trending
                    WHERE to_char(date,'MM-YYYY') in ('01-2020','02-2020','03-2020','04-2020','05-2020','06-2020','07-2020',
                                                '08-2020','09-2020','10-2020','11-2020','12-2020')
                    GROUP BY
                        keyword
                      
                    ORDER BY sum DESC
                    Fetch first 5 rows only;''',conn)
        
        sql_620 = {}
        for i in range(0,5):
            sql_620[i] = pd.read_sql_query(''' 
                    SELECT month_year 
                    FROM
                    (SELECT
                        to_char(date,'MM-YYYY') as month_year,
                        (SUM(value)) as sum,
                        keyword
                    FROM vn_trending
                    WHERE keyword in ('{}')
                    GROUP BY
                        keyword,
                        month_year
                    ORDER BY sum DESC) as foo
                    WHERE month_year in ('01-2020','02-2020','03-2020','04-2020','05-2020','06-2020','07-2020',
                                                '08-2020','09-2020','10-2020','11-2020','12-2020')
                    FETCH first 1 rows only;'''.format(df5_20.iloc[i,0]),conn)
            top_month20 = pd.concat(sql_620,ignore_index=True)      
            top_keytrend20 = pd.concat([df5_20,top_month20],axis=1)
            top_keytrend20.insert(0,'STT',[i for i in range(1,len(df5_20)+1)],True)        
            top_keytrend20.columns = ['STT','keyword','Số lần tìm kiếm','Tháng tìm kiếm nhiều nhất']       
            
        #top 5 tu khoa nhieu nhat 2019
        import psycopg2
        import pandas as pd
        import numpy as np
        conn = psycopg2.connect("host=localhost dbname=postgres user=postgres password=1234")

        #truy vấn dữ liệu năm 2019
        df5_19 = pd.read_sql_query('''
                    SELECT
                        keyword,
                        (SUM(value)) as sum
                        
                    FROM vn_trending
                    WHERE to_char(date,'MM-YYYY') in ('01-2019','02-2019','03-2019','04-2019','05-2019','06-2019','07-2019',
                                                '08-2019','09-2019','10-2019','11-2019','12-2019') 
                    GROUP BY
                        keyword
                      
                    ORDER BY sum DESC
                    Fetch first 5 rows only;''',conn)
        
        sql_619 = {}
        for i in range(0,5):
            sql_619[i] = pd.read_sql_query(''' 
                    SELECT month_year 
                    FROM
                    (SELECT
                        to_char(date,'MM-YYYY') as month_year,
                        (SUM(value)) as sum,
                        keyword
                    FROM vn_trending
                    WHERE keyword in ('{}')
                    GROUP BY
                        keyword,
                        month_year
                    ORDER BY sum DESC) as foo
                    WHERE month_year in ('01-2019','02-2019','03-2019','04-2019','05-2019','06-2019','07-2019',
                                                '08-2019','09-2019','10-2019','11-2019','12-2019') 
                    GROUP BY
                        month_year,
                        keyword
                    FETCH first 1 rows only;'''.format(df5_19.iloc[i,0]), conn)
            top_month19 = pd.concat(sql_619,ignore_index=True)      
            top_keytrend19 = pd.concat([df5_19,top_month19],axis=1)
            top_keytrend19.columns = ['keyword','Số lần tìm kiếm','Tháng tìm kiếm nhiều nhất']       
        # concat 2 bảng data lại
        thongke = pd.concat([top_keytrend20,top_keytrend19], axis=1)
        
         #lưu  thong ke vào file merge đã tạo
        import pandas as pd
        import openpyxl
        from openpyxl import load_workbook
        writer = pd.ExcelWriter('result_user_6.xlsx', engine='openpyxl')
            # try to open an existing workbook
        writer.book = load_workbook('result_user_6.xlsx')
            # copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            # read existing file
        reader = pd.read_excel(r'result_user_6.xlsx')
            # write out the new sheet
        thongke.to_excel(writer,index=False,header=True,startrow=len(reader)+1)
        writer.close()
        
        print('Đã xuất file result_user_6.xlsx thành công!!!')
        continue
    elif user == '99':
        break
print('Đã thoát chương trình!!!')
    
    
        

